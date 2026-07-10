#!/usr/bin/env python3
"""Informe Excel de cobertura línea base ↔ línea salida, por ciudad y colegio.

Reusa el motor de `lib/coverage.py` (mismo cruce que el dashboard y `cruce_cobertura.py`)
y produce un .xlsx multi-hoja en `outputs/` (gitignored — lleva PII de los faltantes):

  1. Resumen          — KPIs globales + cobertura por ciudad/colegio + métodos de match.
  2. Cobertura x grado — detalle por ciudad/colegio/grado (con cohorte en Iquitos).
  3. Faltan           — nominal de la base Tratamiento no re-alcanzada ("de menos").
  4. Revisar match    — matches por nombre-fuzzy o teléfono (baja confianza) a verificar
                        a mano: base vs endline lado a lado.
  5. De más           — endline en colegios Tratamiento sin match en base (nuevos / IDs
                        sucios) a revisar a mano.

Uso:
    cd ~/Documents/Dev/AMA/Lineasalida2026/lago-agrio
    .venv/bin/python scripts/informe_cobertura_excel.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "dashboard"))
os.chdir(ROOT)

import streamlit as st  # noqa: E402  (solo para credenciales/secrets)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.utils.dataframe import dataframe_to_rows  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402

from lib import coverage as cov  # noqa: E402
from lib import db, kobo  # noqa: E402

CIUDAD_DISPLAY = {"iquitos": "Iquitos", "lago_agrio": "Lago Agrio"}
METODO_DISPLAY = {
    "documento": "Documento (alta confianza)",
    "nombre+colegio": "Nombre+colegio exacto (alta)",
    "nombre+colegio~fuzzy": "Nombre fuzzy (REVISAR)",
    "telefono": "Teléfono (REVISAR)",
}

# ── Estilos ──────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")     # teal oscuro
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F4E5F")
SUB_FONT = Font(italic=True, size=10, color="555555")
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
REVISAR_FILL = PatternFill("solid", fgColor="FFF2CC")


def pct_fill(p):
    if p is None:
        return None
    if p >= 70:
        return GREEN
    if p >= 40:
        return YELLOW
    return RED


def cohorte(grado: str) -> str:
    g = str(grado).strip()
    if g.startswith("4"):
        return "Escolar (4°→5°)"
    if g.startswith("5"):
        return "Egresado (5°→egresa)"
    return "—"


def fetch_endline() -> pd.DataFrame:
    frames = []
    try:
        frames.append(kobo.get_identities(st.secrets["KOBO_ASSET_UID"]))
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ Kobo falló: {e}")
    try:
        frames.append(db.get_identities(st.secrets["FORM_ID"]))
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ Typeform falló: {e}")
    return cov.build_endline_identities(*frames)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_df(ws, df, start_row=1, title=None, subtitle=None, autofilter=True):
    """Escribe un DataFrame con header estilizado. Devuelve la fila siguiente libre."""
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 1
    if subtitle:
        ws.cell(row=r, column=1, value=subtitle).font = SUB_FONT
        r += 1
    if title or subtitle:
        r += 1
    hdr_row = r
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=hdr_row, column=j, value=str(col))
    style_header(ws, hdr_row, len(df.columns))
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, col in enumerate(df.columns, 1):
            v = row[col]
            if isinstance(v, float) and pd.isna(v):
                v = ""
            ws.cell(row=hdr_row + i, column=j, value=v).border = BORDER
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    if autofilter:
        ws.auto_filter.ref = (
            f"A{hdr_row}:{get_column_letter(len(df.columns))}{hdr_row + len(df)}"
        )
    return hdr_row, hdr_row + len(df)


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    stamp = datetime.now().strftime("%Y-%m-%d")
    print("Cargando base…")
    base = cov.load_baseline(
        str(ROOT.parent.parent / "Preprocesamiento" / "outputs" / "AMA_encuesta_unificada.csv")
    )
    print(f"  Tratamiento: {len(base)}")
    print("Trayendo endline…")
    end = fetch_endline()
    print(f"  Endline único: {len(end)}")
    res = cov.match(base, end)
    mb = res.baseline.copy()
    mb["ciudad_d"] = mb["ciudad"].map(lambda c: CIUDAD_DISPLAY.get(c, c))

    # endline lookup por response_id para el lado-a-lado
    elook = end.set_index("response_id")

    # de más solo en colegios Tratamiento
    trat = set(base["ncolegio"].dropna())
    nb = res.endline_no_en_base
    demas = nb[nb["ncolegio"].isin(trat)].copy() if not nb.empty else nb.copy()
    demas["ciudad_d"] = demas["ciudad"].map(lambda c: CIUDAD_DISPLAY.get(c, c))

    from openpyxl import Workbook
    wb = Workbook()

    # ── HOJA 1 · RESUMEN ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    tot_meta, tot_alc = len(mb), int(mb["alcanzado"].sum())
    ws.cell(row=1, column=1, value="Informe de cobertura · Línea de salida vs línea base — AMA").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"Generado {stamp}  ·  Meta (base Tratamiento) = {tot_meta}  ·  "
                  f"Alcanzados = {tot_alc} ({tot_alc/tot_meta*100:.1f}%)  ·  "
                  f"Faltan = {tot_meta-tot_alc}  ·  De más (a revisar) = {len(demas)}").font = SUB_FONT

    rows = []
    for (cd, col), sub in mb.groupby(["ciudad_d", "colegio"], dropna=False):
        a, m = int(sub["alcanzado"].sum()), len(sub)
        vc = sub["metodo"].value_counts()
        nb_col = len(demas[demas["colegio"] == col])
        rows.append({
            "Ciudad": cd, "Colegio": col, "Meta (base)": m,
            "Alcanzados": a, "Faltan": m - a,
            "% Cobertura": round(a / m * 100, 1) if m else 0,
            "De más (revisar)": nb_col,
            "↳ Documento": int(vc.get("documento", 0)),
            "↳ Nombre exacto": int(vc.get("nombre+colegio", 0)),
            "↳ Fuzzy (revisar)": int(vc.get("nombre+colegio~fuzzy", 0)),
            "↳ Teléfono (revisar)": int(vc.get("telefono", 0)),
        })
    resumen = pd.DataFrame(rows).sort_values(["Ciudad", "Colegio"]).reset_index(drop=True)
    # fila total
    tot = {
        "Ciudad": "TOTAL", "Colegio": "", "Meta (base)": resumen["Meta (base)"].sum(),
        "Alcanzados": resumen["Alcanzados"].sum(), "Faltan": resumen["Faltan"].sum(),
        "% Cobertura": round(resumen["Alcanzados"].sum() / resumen["Meta (base)"].sum() * 100, 1),
        "De más (revisar)": resumen["De más (revisar)"].sum(),
        "↳ Documento": resumen["↳ Documento"].sum(),
        "↳ Nombre exacto": resumen["↳ Nombre exacto"].sum(),
        "↳ Fuzzy (revisar)": resumen["↳ Fuzzy (revisar)"].sum(),
        "↳ Teléfono (revisar)": resumen["↳ Teléfono (revisar)"].sum(),
    }
    resumen_w = pd.concat([resumen, pd.DataFrame([tot])], ignore_index=True)
    h0, h1 = write_df(ws, resumen_w, start_row=4)
    # color en % cobertura y negrita en total
    pct_idx = list(resumen_w.columns).index("% Cobertura") + 1
    for i in range(len(resumen_w)):
        rr = h0 + 1 + i
        cell = ws.cell(row=rr, column=pct_idx)
        f = pct_fill(resumen_w.iloc[i]["% Cobertura"])
        if f:
            cell.fill = f
        cell.alignment = CENTER
        if resumen_w.iloc[i]["Ciudad"] == "TOTAL":
            for c in range(1, len(resumen_w.columns) + 1):
                ws.cell(row=rr, column=c).font = TOTAL_FONT
    autosize(ws, {"A": 12, "B": 34, "C": 12, "D": 12, "E": 9, "F": 13,
                  "G": 16, "H": 13, "I": 15, "J": 17, "K": 19})

    # leyenda
    lr = h1 + 3
    ws.cell(row=lr, column=1, value="Cómo leer este informe").font = Font(bold=True, size=12, color="1F4E5F")
    notas = [
        "• Meta = personas medidas en la línea base, grupo Tratamiento (universo de seguimiento).",
        "• Alcanzados = re-encontrados en la salida. Faltan = 'de menos' (hoja 'Faltan').",
        "• Métodos de match, de mayor a menor confianza: Documento ▸ Nombre+colegio exacto ▸ "
        "Nombre fuzzy ▸ Teléfono.",
        "• 'Fuzzy' y 'Teléfono' son de baja confianza → hoja 'Revisar match' (verificar a mano).",
        "• 'De más' = encuestados en colegios Tratamiento sin match en base: posibles estudiantes "
        "nuevos o documentos mal digitados → hoja 'De más'.",
        "• Iquitos: la hoja por grado separa cohorte Escolar (4°→5°) de Egresado (5°→egresa).",
        "• Contiene PII (nombre/documento/teléfono): archivo de uso interno, no publicar.",
    ]
    for i, n in enumerate(notas):
        ws.cell(row=lr + 1 + i, column=1, value=n).font = Font(size=10)

    # ── HOJA 2 · COBERTURA x GRADO ───────────────────────────────────────────
    ws2 = wb.create_sheet("Cobertura x grado")
    agg = res.aggregate.copy()
    agg["Ciudad"] = agg["ciudad"].map(lambda c: CIUDAD_DISPLAY.get(c, c))
    agg["Cohorte"] = agg.apply(
        lambda r: cohorte(r["grado"]) if r["ciudad"] == "iquitos" else "—", axis=1
    )
    agg = agg.rename(columns={"colegio": "Colegio", "grado": "Grado", "meta": "Meta",
                              "alcanzados": "Alcanzados", "faltan": "Faltan", "pct": "% Cobertura"})
    agg = agg[["Ciudad", "Colegio", "Grado", "Cohorte", "Meta", "Alcanzados", "Faltan", "% Cobertura"]]
    agg = agg.sort_values(["Ciudad", "Colegio", "Grado"]).reset_index(drop=True)
    g0, g1 = write_df(ws2, agg, title="Cobertura por ciudad / colegio / grado",
                      subtitle=f"Generado {stamp}")
    pidx = list(agg.columns).index("% Cobertura") + 1
    for i in range(len(agg)):
        cell = ws2.cell(row=g0 + 1 + i, column=pidx)
        f = pct_fill(agg.iloc[i]["% Cobertura"])
        if f:
            cell.fill = f
        cell.alignment = CENTER
    autosize(ws2, {"A": 12, "B": 34, "C": 14, "D": 20, "E": 8, "F": 11, "G": 8, "H": 13})

    # ── HOJA 3 · FALTAN (de menos) ───────────────────────────────────────────
    ws3 = wb.create_sheet("Faltan (de menos)")
    fal = mb[~mb["alcanzado"]].copy()
    fal["Ciudad"] = fal["ciudad"].map(lambda c: CIUDAD_DISPLAY.get(c, c))
    fal["Cohorte"] = fal.apply(
        lambda r: cohorte(r["grado"]) if r["ciudad"] == "iquitos" else "—", axis=1)
    fal["Tiene documento"] = fal["ndoc"].map(lambda v: "sí" if isinstance(v, str) and v else "NO")
    fal = fal[["Ciudad", "colegio", "grado", "Cohorte", "codigo", "nombre",
               "documento", "telefono", "Tiene documento"]].rename(
        columns={"colegio": "Colegio", "grado": "Grado", "codigo": "Código AMA",
                 "nombre": "Nombre", "documento": "Documento", "telefono": "Teléfono"})
    fal = fal.sort_values(["Ciudad", "Colegio", "Grado", "Nombre"]).reset_index(drop=True)
    write_df(ws3, fal, title=f"Faltan — base Tratamiento no re-alcanzada ({len(fal)})",
             subtitle=f"Generado {stamp} · PII · uso interno · 'Tiene documento'=NO ⇒ re-buscar por nombre")
    autosize(ws3, {"A": 12, "B": 34, "C": 12, "D": 20, "E": 12, "F": 30,
                   "G": 16, "H": 14, "I": 15})

    # ── HOJA 4 · REVISAR MATCH (fuzzy + teléfono) ────────────────────────────
    ws4 = wb.create_sheet("Revisar match")
    dud = mb[mb["metodo"].isin(["nombre+colegio~fuzzy", "telefono"])].copy()
    out_rows = []
    for _, b in dud.iterrows():
        rid = b["endline_response_id"]
        e = elook.loc[rid] if rid in elook.index else None
        if isinstance(e, pd.DataFrame):  # rid duplicado defensivo
            e = e.iloc[0]
        out_rows.append({
            "Ciudad": CIUDAD_DISPLAY.get(b["ciudad"], b["ciudad"]),
            "Colegio": b["colegio"], "Grado": b["grado"],
            "Método": "Nombre fuzzy" if b["metodo"].endswith("fuzzy") else "Teléfono",
            "Código AMA": b["codigo"],
            "BASE · Nombre": b["nombre"], "BASE · Documento": b["documento"],
            "BASE · Teléfono": b["telefono"],
            "SALIDA · Nombre": (e["nombre"] if e is not None else ""),
            "SALIDA · Documento": (e["documento"] if e is not None else ""),
            "SALIDA · Teléfono": (e["telefono"] if e is not None else ""),
            "SALIDA · Colegio": (e["colegio"] if e is not None else ""),
            "Fuente": (e["fuente"] if e is not None else ""),
            "¿Es la misma persona? (sí/no)": "",
        })
    dudf = pd.DataFrame(out_rows).sort_values(["Ciudad", "Colegio", "Método", "BASE · Nombre"]).reset_index(drop=True)
    d0, d1 = write_df(
        ws4, dudf,
        title=f"Matches de baja confianza a verificar a mano ({len(dudf)})",
        subtitle=f"Generado {stamp} · compará BASE vs SALIDA; el teléfono puede ser compartido (padre/madre)")
    # tinte de la fila + columna de decisión
    dec_idx = len(dudf.columns)
    for i in range(len(dudf)):
        for c in range(1, len(dudf.columns) + 1):
            ws4.cell(row=d0 + 1 + i, column=c).fill = REVISAR_FILL
    autosize(ws4, {"A": 11, "B": 30, "C": 12, "D": 13, "E": 12, "F": 28, "G": 15,
                   "H": 13, "I": 28, "J": 15, "K": 13, "L": 28, "M": 10, "N": 26})

    # ── HOJA 5 · DE MÁS ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("De más")
    dm = demas.copy()
    dm["¿Nuevo o ID sucio? (revisar)"] = ""
    dm = dm[["ciudad_d", "colegio", "grado", "nombre", "documento", "telefono",
             "fuente", "response_id", "¿Nuevo o ID sucio? (revisar)"]].rename(
        columns={"ciudad_d": "Ciudad", "colegio": "Colegio", "grado": "Grado",
                 "nombre": "Nombre", "documento": "Documento", "telefono": "Teléfono",
                 "fuente": "Fuente", "response_id": "ID respuesta"})
    dm = dm.sort_values(["Ciudad", "Colegio", "Nombre"]).reset_index(drop=True)
    write_df(ws5, dm,
             title=f"De más — encuestados en colegios Tratamiento sin match en base ({len(dm)})",
             subtitle=f"Generado {stamp} · posibles estudiantes nuevos o documentos mal digitados")
    autosize(ws5, {"A": 11, "B": 32, "C": 14, "D": 30, "E": 16, "F": 14,
                   "G": 11, "H": 24, "I": 26})

    out_path = ROOT / "outputs" / f"informe_cobertura_AMA_{stamp}.xlsx"
    wb.save(out_path)
    print(f"\n✅ Informe → {out_path}")
    print(f"   Hojas: Resumen · Cobertura x grado · Faltan ({len(fal)}) · "
          f"Revisar match ({len(dudf)}) · De más ({len(dm)})")


if __name__ == "__main__":
    main()
