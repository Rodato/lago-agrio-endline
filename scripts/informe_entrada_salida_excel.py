#!/usr/bin/env python3
"""Export entrada + salida en un solo .xlsx, una pestaña por encuesta.

  · Entrada — consolidado de la línea base (Lago Agrio + Iquitos, todos los grupos),
              a nivel persona.
  · Salida  — identidades de la línea de salida ya cruzadas (Kobo + Typeform) y limpias
              (deduplicadas, con fix de placeholders), a nivel persona.

Reusa `lib/coverage.py` para cargar/normalizar/deduplicar (mismo motor que cobertura).
Salida: `outputs/entrada_salida_AMA_<fecha>.xlsx` (gitignored — **lleva PII de menores**,
uso interno, NO publicar/deployar).

Uso:
    cd ~/Documents/Dev/AMA/Lineasalida2026/lago-agrio
    .venv/bin/python scripts/informe_entrada_salida_excel.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "dashboard"))
os.chdir(ROOT)

import streamlit as st  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from lib import coverage as cov  # noqa: E402
from lib import db, kobo  # noqa: E402

BASE_PATH = str(ROOT.parent.parent / "Preprocesamiento" / "outputs" / "AMA_encuesta_unificada.csv")
CIUDAD_DISPLAY = {"iquitos": "Iquitos", "lago_agrio": "Lago Agrio"}

HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E5F")
SUB_FONT = Font(italic=True, size=10, color="555555")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _ciudad(code):
    return CIUDAD_DISPLAY.get(code, "(sin ciudad)")


def write_sheet(ws, df, title, subtitle, widths, center_cols=()):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    hdr = 4
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=hdr, column=j, value=str(col))
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, col in enumerate(df.columns, 1):
            v = row[col]
            if isinstance(v, float) and pd.isna(v):
                v = ""
            cell = ws.cell(row=hdr + i, column=j, value=v)
            cell.border = BORDER
            if col in center_cols:
                cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(df.columns))}{hdr + len(df)}"
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    stamp = datetime.now().strftime("%Y-%m-%d")

    # ── ENTRADA: consolidado base, todos los grupos ──────────────────────────────
    print("Cargando entrada (consolidado base)…")
    base = cov.load_baseline(BASE_PATH, grupo=None)
    ent = base.copy()
    ent["Ciudad"] = ent["ciudad"].map(_ciudad)
    ent = ent.rename(columns={
        "colegio": "Colegio", "grado": "Grado", "grupo": "Grupo", "codigo": "Código AMA",
        "nombre": "Nombre", "documento": "Documento", "telefono": "Teléfono"})
    ent = ent[["Ciudad", "Colegio", "Grado", "Grupo", "Código AMA",
               "Nombre", "Documento", "Teléfono"]]
    ent = ent.sort_values(["Ciudad", "Colegio", "Grado", "Nombre"]).reset_index(drop=True)
    print(f"  Entrada: {len(ent)} personas")

    # ── SALIDA: identidades cruzadas (Kobo+Typeform) y limpias ───────────────────
    print("Trayendo salida (Kobo + Typeform)…")
    end = cov.build_endline_identities(
        kobo.get_identities(st.secrets["KOBO_ASSET_UID"]),
        db.get_identities(st.secrets["FORM_ID"]),
    )
    sal = end.copy()
    sal["Ciudad"] = sal["cciudad"].map(_ciudad)
    sal["Fecha envío"] = pd.to_datetime(sal["submitted_at"], utc=True, errors="coerce") \
        .dt.tz_convert(None).dt.strftime("%Y-%m-%d %H:%M")
    sal["Fuente"] = sal["fuente"].map({"kobo": "Kobo", "typeform": "Typeform"}).fillna(sal["fuente"])
    sal = sal.rename(columns={
        "colegio": "Colegio", "grado": "Grado", "nombre": "Nombre", "documento": "Documento",
        "telefono": "Teléfono", "correo": "Correo", "response_id": "ID respuesta"})
    sal = sal[["Ciudad", "Colegio", "Grado", "Fuente", "Nombre", "Documento",
               "Teléfono", "Correo", "Fecha envío", "ID respuesta"]]
    sal = sal.sort_values(["Ciudad", "Colegio", "Grado", "Nombre"]).reset_index(drop=True)
    print(f"  Salida: {len(sal)} personas únicas")

    # ── Excel ────────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Entrada"
    write_sheet(
        ws1, ent,
        "Línea de entrada · consolidado AMA (Lago Agrio + Iquitos)",
        f"Generado {stamp} · {len(ent)} personas (todos los grupos) · PII · uso interno, no publicar",
        widths={"A": 12, "B": 34, "C": 14, "D": 13, "E": 14, "F": 32, "G": 16, "H": 15},
        center_cols={"Ciudad", "Grado", "Grupo"},
    )

    ws2 = wb.create_sheet("Salida")
    write_sheet(
        ws2, sal,
        "Línea de salida · identidades cruzadas y limpias (Kobo + Typeform)",
        f"Generado {stamp} · {len(sal)} personas únicas (deduplicado, fix placeholders) · "
        f"PII · uso interno, no publicar",
        widths={"A": 12, "B": 34, "C": 14, "D": 11, "E": 32, "F": 16, "G": 15,
                "H": 26, "I": 17, "J": 38},
        center_cols={"Ciudad", "Grado", "Fuente"},
    )

    out = ROOT / "outputs" / f"entrada_salida_AMA_{stamp}.xlsx"
    wb.save(out)
    print(f"\n✅ → {out}")
    print(f"   Entrada {len(ent)} · Salida {len(sal)}")


if __name__ == "__main__":
    main()
