#!/usr/bin/env python3
"""Informe de línea de salida (endline) — encuestados por ciudad y colegio.

A diferencia del informe de cobertura (que cruza contra la base Tratamiento), este es
un conteo simple de la encuesta de salida: cuántas **personas únicas** respondió cada
colegio, en ambas ciudades, sobre **todos** los colegios encuestados (Tratamiento +
Control + otros). Reusa el motor de `lib/coverage.py` para la deduplicación por persona
(incluye el filtro de documentos placeholder: 11111111, 12345678, etc.).

Salida: `outputs/informe_salida_AMA_<fecha>.xlsx` (gitignored — lleva conteos, sin PII):
  · Resumen por ciudad  +  detalle por colegio, con desglose Kobo / Typeform.

Uso:
    cd ~/Documents/Dev/AMA/Lineasalida2026/lago-agrio
    .venv/bin/python scripts/informe_salida_excel.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "dashboard"))
os.chdir(ROOT)

import streamlit as st  # noqa: E402  (solo para credenciales/secrets)
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from lib import coverage as cov  # noqa: E402
from lib import db, kobo  # noqa: E402

CIUDAD_DISPLAY = {"iquitos": "Iquitos", "lago_agrio": "Lago Agrio"}
SIN_CIUDAD = "(sin ciudad)"
SIN_COLEGIO = "(colegio sin registrar)"

# ── Estilos (mismos que el informe de cobertura) ─────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F4E5F")
SUB_FONT = Font(italic=True, size=10, color="555555")
TOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor="DCE6EB")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fetch_endline() -> pd.DataFrame:
    frames = []
    try:
        frames.append(kobo.get_identities(st.secrets["KOBO_ASSET_UID"]))
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ Kobo falló: {e}")
    try:
        frames.append(db.get_identities(st.secrets["FORM_ID"]))
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ Typeform falló: {e}")
    return cov.build_endline_identities(*frames)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, df, hdr_row, center_from=3):
    """Escribe df con header estilizado en hdr_row. Devuelve la fila siguiente libre."""
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=hdr_row, column=j, value=str(col))
    style_header(ws, hdr_row, len(df.columns))
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, col in enumerate(df.columns, 1):
            v = row[col]
            if isinstance(v, float) and pd.isna(v):
                v = ""
            cell = ws.cell(row=hdr_row + i, column=j, value=v)
            cell.border = BORDER
            if j >= center_from:
                cell.alignment = CENTER
    return hdr_row + len(df)


def main():
    stamp = datetime.now().strftime("%Y-%m-%d")
    print("Trayendo línea de salida (Kobo + Typeform)…")
    end = fetch_endline()
    print(f"  Personas únicas: {len(end)}")

    end = end.copy()
    end["Ciudad"] = end["cciudad"].map(CIUDAD_DISPLAY).fillna(SIN_CIUDAD)
    end["Colegio"] = end["colegio"].where(end["colegio"].map(lambda v: isinstance(v, str) and bool(v)), SIN_COLEGIO)
    end["Fuente"] = end["fuente"].fillna("?")

    # ── Detalle por colegio ──────────────────────────────────────────────────────
    g = end.groupby(["Ciudad", "Colegio"])
    detalle = g.agg(
        Encuestados=("response_id", "size"),
        Kobo=("Fuente", lambda s: int((s == "kobo").sum())),
        Typeform=("Fuente", lambda s: int((s == "typeform").sum())),
    ).reset_index()
    detalle = detalle.sort_values(["Ciudad", "Encuestados"], ascending=[True, False]).reset_index(drop=True)

    # ── Resumen por ciudad ───────────────────────────────────────────────────────
    total = len(end)
    resumen = end.groupby("Ciudad").agg(
        Encuestados=("response_id", "size"),
        Kobo=("Fuente", lambda s: int((s == "kobo").sum())),
        Typeform=("Fuente", lambda s: int((s == "typeform").sum())),
        Colegios=("Colegio", "nunique"),
    ).reset_index().sort_values("Encuestados", ascending=False)
    resumen["% del total"] = (resumen["Encuestados"] / total * 100).round(1)
    resumen = resumen[["Ciudad", "Colegios", "Encuestados", "% del total", "Kobo", "Typeform"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Salida por colegio"

    ws.cell(row=1, column=1,
            value="Informe de línea de salida · AMA — encuestados por ciudad y colegio").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=(f"Generado {stamp}  ·  Encuestados únicos = {total}  ·  "
                   f"Iquitos = {int(resumen.loc[resumen['Ciudad']=='Iquitos','Encuestados'].sum())}  ·  "
                   f"Lago Agrio = {int(resumen.loc[resumen['Ciudad']=='Lago Agrio','Encuestados'].sum())}  ·  "
                   f"Fuentes: Kobo + Typeform")).font = SUB_FONT

    # Resumen por ciudad
    ws.cell(row=4, column=1, value="Resumen por ciudad").font = Font(bold=True, size=12, color="1F4E5F")
    res_hdr = 5
    res_end = write_table(ws, resumen, res_hdr, center_from=2)
    # fila total del resumen
    rt = res_end + 1
    ws.cell(row=rt, column=1, value="TOTAL")
    ws.cell(row=rt, column=2, value=int(resumen["Colegios"].sum()))
    ws.cell(row=rt, column=3, value=total)
    ws.cell(row=rt, column=4, value=100.0)
    ws.cell(row=rt, column=5, value=int(resumen["Kobo"].sum()))
    ws.cell(row=rt, column=6, value=int(resumen["Typeform"].sum()))
    for c in range(1, 7):
        cell = ws.cell(row=rt, column=c)
        cell.font = TOTAL_FONT
        cell.border = BORDER
        if c >= 2:
            cell.alignment = CENTER

    # Detalle por colegio (con subtotal por ciudad + total general)
    det_title = rt + 3
    ws.cell(row=det_title, column=1, value="Detalle por colegio").font = Font(bold=True, size=12, color="1F4E5F")
    det_hdr = det_title + 1
    # columnas a mostrar
    cols = ["Ciudad", "Colegio", "Encuestados", "Kobo", "Typeform"]
    for j, col in enumerate(cols, 1):
        ws.cell(row=det_hdr, column=j, value=col)
    style_header(ws, det_hdr, len(cols))
    r = det_hdr + 1
    for ciudad in resumen["Ciudad"]:
        sub = detalle[detalle["Ciudad"] == ciudad]
        for _, row in sub.iterrows():
            for j, col in enumerate(cols, 1):
                cell = ws.cell(row=r, column=j, value=row[col])
                cell.border = BORDER
                if j >= 3:
                    cell.alignment = CENTER
            r += 1
        # subtotal ciudad
        ws.cell(row=r, column=1, value=f"Subtotal {ciudad}")
        ws.cell(row=r, column=3, value=int(sub["Encuestados"].sum()))
        ws.cell(row=r, column=4, value=int(sub["Kobo"].sum()))
        ws.cell(row=r, column=5, value=int(sub["Typeform"].sum()))
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = TOTAL_FONT
            cell.fill = SUBTOTAL_FILL
            cell.border = BORDER
            if c >= 3:
                cell.alignment = CENTER
        r += 1
    # total general
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=3, value=total)
    ws.cell(row=r, column=4, value=int(detalle["Kobo"].sum()))
    ws.cell(row=r, column=5, value=int(detalle["Typeform"].sum()))
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.font = TOTAL_FONT
        cell.border = BORDER
        if c >= 3:
            cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=det_hdr + 1, column=1)

    # notas
    nr = r + 2
    ws.cell(row=nr, column=1, value="Notas").font = Font(bold=True, size=11, color="1F4E5F")
    notas = [
        "• Encuestados = personas únicas (deduplicadas por documento, o nombre+colegio si no hay documento).",
        "• Incluye todos los colegios encuestados en la salida: Tratamiento, Control y otros.",
        f"• '{SIN_COLEGIO}' = el envío no trae colegio registrado (no es un colegio nuevo).",
        "• Kobo / Typeform = fuente del registro que sobrevive al deduplicar (en colegios con ambos "
        "canales una misma persona cuenta una vez).",
        "• Documentos placeholder (11111111, 12345678…) se tratan como 'sin documento' para no "
        "colapsar personas distintas que comparten el relleno.",
    ]
    for i, n in enumerate(notas):
        ws.cell(row=nr + 1 + i, column=1, value=n).font = Font(size=10)

    # anchos
    widths = {"A": 20, "B": 34, "C": 14, "D": 12, "E": 12, "F": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out_path = ROOT / "outputs" / f"informe_salida_AMA_{stamp}.xlsx"
    wb.save(out_path)
    print(f"\n✅ Informe → {out_path}")
    print(f"   Iquitos {int(resumen.loc[resumen['Ciudad']=='Iquitos','Encuestados'].sum())} · "
          f"Lago Agrio {int(resumen.loc[resumen['Ciudad']=='Lago Agrio','Encuestados'].sum())} · "
          f"TOTAL {total} encuestados únicos en {len(detalle)} colegios")


if __name__ == "__main__":
    main()
