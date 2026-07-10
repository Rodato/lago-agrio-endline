#!/usr/bin/env python3
"""Comparativo línea base ↔ línea salida — totales por ciudad y colegio.

Pone lado a lado, para cada ciudad y colegio, cuánta gente se midió en la **línea base**
(todos los grupos: Tratamiento + Control) vs cuántas **personas únicas** respondió la
**línea de salida** (endline, Kobo + Typeform deduplicado). NO es cobertura (eso cruza
persona a persona); acá son conteos agregados de volumen por colegio.

Salida: `outputs/informe_base_vs_salida_AMA_<fecha>.xlsx` (gitignored — solo conteos):
  · Resumen por ciudad  ·  Resumen por grupo  ·  Detalle por colegio (base, salida, Δ).

Uso:
    cd ~/Documents/Dev/AMA/Lineasalida2026/lago-agrio
    .venv/bin/python scripts/informe_base_vs_salida_excel.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "dashboard"))
os.chdir(ROOT)

import streamlit as st  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402

from lib import coverage as cov  # noqa: E402
from lib import db, kobo  # noqa: E402

BASE_PATH = str(ROOT.parent.parent / "Preprocesamiento" / "outputs" / "AMA_encuesta_unificada.csv")
CIUDAD_DISPLAY = {"iquitos": "Iquitos", "lago_agrio": "Lago Agrio"}
SIN_CIUDAD = "(sin ciudad)"
SIN_COLEGIO = "(colegio sin registrar)"

HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F4E5F")
SECTION_FONT = Font(bold=True, size=12, color="1F4E5F")
SUB_FONT = Font(italic=True, size=10, color="555555")
TOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor="DCE6EB")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _ciudad_d(code):
    return CIUDAD_DISPLAY.get(code, SIN_CIUDAD)


def _colegio_key(ncol, display):
    """Llave de join y nombre a mostrar; colapsa vacío/ilegible a 'sin colegio'."""
    if not isinstance(ncol, str) or not ncol or ncol == "ilegible":
        return ("__sin__", SIN_COLEGIO)
    return (ncol, display if isinstance(display, str) and display else ncol)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER


def main():
    stamp = datetime.now().strftime("%Y-%m-%d")

    print("Cargando base (todos los grupos)…")
    base = cov.load_baseline(BASE_PATH, grupo=None)
    print(f"  Base: {len(base)}")
    print("Trayendo línea de salida…")
    end = cov.build_endline_identities(
        kobo.get_identities(st.secrets["KOBO_ASSET_UID"]),
        db.get_identities(st.secrets["FORM_ID"]),
    )
    print(f"  Salida (únicos): {len(end)}")

    # ── normalizar ambos lados a (ciudad, colegio_key) ───────────────────────────
    def prep(df, city_col):
        d = df.copy()
        d["Ciudad"] = d[city_col].map(_ciudad_d)
        keys = [_colegio_key(nc, disp) for nc, disp in zip(d["ncolegio"], d["colegio"])]
        d["ckey"] = [k for k, _ in keys]
        d["Colegio"] = [name for _, name in keys]
        return d

    b = prep(base, "ciudad")
    e = prep(end, "cciudad")

    b_tab = b.groupby(["Ciudad", "ckey", "Colegio"]).agg(
        Base=("codigo", "size"),
        Grupo=("grupo", lambda s: "/".join(sorted(x for x in s.dropna().unique() if x))),
    ).reset_index()
    e_tab = e.groupby(["Ciudad", "ckey", "Colegio"]).agg(Salida=("response_id", "size")).reset_index()

    det = b_tab.merge(e_tab, on=["Ciudad", "ckey", "Colegio"], how="outer")
    det["Base"] = det["Base"].fillna(0).astype(int)
    det["Salida"] = det["Salida"].fillna(0).astype(int)
    det["Grupo"] = det["Grupo"].replace("", pd.NA).fillna("—")
    det["Δ (salida−base)"] = det["Salida"] - det["Base"]
    det["Salida/Base %"] = det.apply(
        lambda r: round(r["Salida"] / r["Base"] * 100, 1) if r["Base"] else None, axis=1)
    det = det.sort_values(["Ciudad", "Base"], ascending=[True, False]).reset_index(drop=True)

    tot_base, tot_sal = int(det["Base"].sum()), int(det["Salida"].sum())

    # ── resumen por ciudad y por grupo ───────────────────────────────────────────
    res_ciudad = det.groupby("Ciudad").agg(
        Colegios=("Colegio", "nunique"), Base=("Base", "sum"), Salida=("Salida", "sum")
    ).reset_index().sort_values("Base", ascending=False)
    res_ciudad["Δ"] = res_ciudad["Salida"] - res_ciudad["Base"]

    res_grupo = det.groupby("Grupo").agg(Base=("Base", "sum"), Salida=("Salida", "sum")).reset_index()
    res_grupo["Δ"] = res_grupo["Salida"] - res_grupo["Base"]
    orden = {"Tratamiento": 0, "Control": 1}
    res_grupo = res_grupo.sort_values("Grupo", key=lambda s: s.map(lambda g: orden.get(g, 9)))

    # ── escribir Excel ───────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Base vs salida"

    ws.cell(row=1, column=1, value="Comparativo línea base ↔ línea salida · AMA — por ciudad y colegio").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        f"Generado {stamp}  ·  Línea base = {tot_base} (todos los grupos)  ·  "
        f"Línea salida = {tot_sal} (personas únicas)  ·  Δ = {tot_sal - tot_base}")).font = SUB_FONT

    r = 4

    def write_block(title, df, cols, total_row=None):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = SECTION_FONT
        r += 1
        for j, c in enumerate(cols, 1):
            ws.cell(row=r, column=j, value=c)
        style_header(ws, r, len(cols))
        r += 1
        for _, row in df.iterrows():
            for j, c in enumerate(cols, 1):
                v = row[c]
                if isinstance(v, float) and pd.isna(v):
                    v = ""
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = BORDER
                if j >= 2:
                    cell.alignment = CENTER
            r += 1
        if total_row is not None:
            for j, c in enumerate(cols, 1):
                cell = ws.cell(row=r, column=j, value=total_row.get(c, ""))
                cell.border, cell.font = BORDER, TOTAL_FONT
                if j >= 2:
                    cell.alignment = CENTER
            r += 1
        r += 1  # línea en blanco

    # Resumen por ciudad
    write_block("Resumen por ciudad", res_ciudad,
                ["Ciudad", "Colegios", "Base", "Salida", "Δ"],
                {"Ciudad": "TOTAL", "Colegios": int(res_ciudad["Colegios"].sum()),
                 "Base": tot_base, "Salida": tot_sal, "Δ": tot_sal - tot_base})

    # Resumen por grupo
    write_block("Resumen por grupo (asignación en la base)", res_grupo,
                ["Grupo", "Base", "Salida", "Δ"],
                {"Grupo": "TOTAL", "Base": tot_base, "Salida": tot_sal, "Δ": tot_sal - tot_base})

    # Detalle por colegio con subtotal por ciudad
    cols = ["Ciudad", "Colegio", "Grupo", "Base", "Salida", "Δ (salida−base)", "Salida/Base %"]
    ws.cell(row=r, column=1, value="Detalle por colegio").font = SECTION_FONT
    r += 1
    det_hdr = r
    for j, c in enumerate(cols, 1):
        ws.cell(row=r, column=j, value=c)
    style_header(ws, r, len(cols))
    r += 1
    for ciudad in res_ciudad["Ciudad"]:
        sub = det[det["Ciudad"] == ciudad]
        for _, row in sub.iterrows():
            for j, c in enumerate(cols, 1):
                v = row[c]
                if isinstance(v, float) and pd.isna(v):
                    v = ""
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = BORDER
                if j >= 3:
                    cell.alignment = CENTER
            r += 1
        sb, ss = int(sub["Base"].sum()), int(sub["Salida"].sum())
        vals = {"Ciudad": f"Subtotal {ciudad}", "Base": sb, "Salida": ss,
                "Δ (salida−base)": ss - sb, "Salida/Base %": round(ss / sb * 100, 1) if sb else ""}
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=r, column=j, value=vals.get(c, ""))
            cell.border, cell.font, cell.fill = BORDER, TOTAL_FONT, SUBTOTAL_FILL
            if j >= 3:
                cell.alignment = CENTER
        r += 1
    # total general
    tvals = {"Ciudad": "TOTAL", "Base": tot_base, "Salida": tot_sal,
             "Δ (salida−base)": tot_sal - tot_base,
             "Salida/Base %": round(tot_sal / tot_base * 100, 1)}
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=r, column=j, value=tvals.get(c, ""))
        cell.border, cell.font = BORDER, TOTAL_FONT
        if j >= 3:
            cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=det_hdr + 1, column=1)
    r += 2

    # notas
    ws.cell(row=r, column=1, value="Notas").font = Font(bold=True, size=11, color="1F4E5F")
    notas = [
        "• Línea base = personas medidas en la entrada (todos los grupos). Línea salida = personas "
        "únicas en la encuesta de salida (Kobo + Typeform deduplicado, con fix de placeholders).",
        "• Esto es VOLUMEN por colegio, NO cobertura: 'Salida/Base %' puede pasar de 100% porque la "
        "salida incluye estudiantes nuevos que no estaban en la base.",
        "• La cobertura real (a cuántos de la base Tratamiento se re-alcanzó persona-a-persona) está "
        "en el informe de cobertura, no acá.",
        f"• '{SIN_COLEGIO}' = envíos/registros sin colegio (base: 'No asignado'/ilegible).",
        "• Grupo = asignación en la base (Tratamiento/Control); los colegios son de un solo grupo.",
    ]
    for i, n in enumerate(notas):
        ws.cell(row=r + 1 + i, column=1, value=n).font = Font(size=10)

    for col, w in {"A": 22, "B": 34, "C": 14, "D": 12, "E": 12, "F": 16, "G": 14}.items():
        ws.column_dimensions[col].width = w

    out = ROOT / "outputs" / f"informe_base_vs_salida_AMA_{stamp}.xlsx"
    wb.save(out)
    print(f"\n✅ Informe → {out}")
    print(f"   Base {tot_base} · Salida {tot_sal} · Δ {tot_sal - tot_base} · {len(det)} colegios")


if __name__ == "__main__":
    main()
